"""Excel workbook writer with multi-sheet generation, styling, and data type inference."""

import os
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pdf_to_sheet.cleaner import is_technical_data_table
from pdf_to_sheet.models import ExtractionResult


def infer_type(val: str) -> Any:
    """Infer numeric / int / float types from cell string."""
    if val is None:
        return ""
    val_str = str(val).strip()
    if not val_str:
        return ""

    # Int check
    if val_str.isdigit() or (val_str.startswith("-") and val_str[1:].isdigit()):
        return int(val_str)

    # Float check
    try:
        if "." in val_str or "," in val_str:
            clean_float = val_str.replace(",", ".")
            return float(clean_float)
    except ValueError:
        pass

    return val_str


class ExcelWriter:
    """Exporter that converts ExtractionResult into formatted XLSX files using openpyxl."""

    def write(self, result: ExtractionResult, output_path: str) -> str:
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default blank sheet

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        all_rows: list[list[Any]] = []
        master_headers: list[str] = []

        # 1. Individual Table Sheets
        for idx, table in enumerate(result.tables, start=1):
            is_tech = is_technical_data_table(table)
            sheet_title = f"Table_{idx}_Page_{table.page_number}"[:31] if is_tech else f"Metadata_Page_{table.page_number}"[:31]
            ws = wb.create_sheet(title=sheet_title)

            if table.headers:
                ws.append(table.headers)
                if is_tech and not master_headers:
                    master_headers = list(table.headers)

            for row in table.rows:
                typed_row = [infer_type(cell) for cell in row]
                ws.append(typed_row)
                if is_tech:
                    all_rows.append(typed_row)

            # Apply formatting
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

            for row_idx, row_cells in enumerate(ws.iter_rows(), start=1):
                for cell in row_cells:
                    cell.border = thin_border
                    if row_idx == 1 and table.headers:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        # 2. Master Consolidated Sheet
        if all_rows:
            ws_master = wb.create_sheet(title="Master_Consolidated", index=0)
            if master_headers:
                ws_master.append(master_headers)
            for row in all_rows:
                ws_master.append(row)

            for col in ws_master.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws_master.column_dimensions[col_letter].width = max(max_len + 4, 12)

            for row_idx, row_cells in enumerate(ws_master.iter_rows(), start=1):
                for cell in row_cells:
                    cell.border = thin_border
                    if row_idx == 1 and master_headers:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        wb.save(output_path)
        return output_path
